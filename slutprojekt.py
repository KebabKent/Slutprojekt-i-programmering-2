import requests
import openpyxl
from getpwd import getpwd
import time

# Jag har kommenterat på en nivå så att de som redan är 
# bra på proggramering vet vad det är som händer 
# med vissa generiska klass funktioner
# och utan att upprepa samma kommentar på varje grej

""" Klasser för anställda """
class Bank:
    def __init__(self, name, password, role, row):
        self.name = name
        self.__password = password
        self.role = role
        self.row = row
    
    def return_password(self):
        return self.__password
    
    def return_row(self):
        return self.row
    
    def return_role(self):
        return self.role

class Admin(Bank): # Admin klassen behövs egentligen inte men skulle se konstigt ut att inte ha en
    def __init__(self, name, password, role, row):
        super().__init__(name, password, role, row)

class User(Bank):
    def __init__(self, name, password, role, row, balance):
        super().__init__(name, password, role, row)
        self.balance = balance

    def check_account_balance(self):
        return self.balance

class Underage_user(Bank):
    def __init__(self,name, password, role, row, balance):
        super().__init__(name, password, role, row)
        self.balance = balance

    def check_account_balance(self):
        return self.balance



""" Loggar in användaren """
def login():
    path = "./Employees.xlsx"
    wb_obj = openpyxl.load_workbook(path) # Öppnar excel
    sheet_obj = wb_obj.active
    max_ro = sheet_obj.max_row # Högsta rad nummret 

    """ Kollar om användaren finns och vart den finns"""
    print('\n<--Loggin-->')
    for i in range(3):
        user_name = str(input('Name: '))
        user_surname = str(input('Surname: '))

        for xx in range(1, max_ro +1):
            user_n = sheet_obj.cell(row = xx, column = 1) # Ger informationen som står på rad xx, column 1
            user_surn = sheet_obj.cell(row = xx, column = 2)

            if user_n.value == user_name and user_surn.value == user_surname:
                print('User exists!\n')
                time.sleep(0.5)

                """ Tar lösenordet och loggar in användaren med användarnamnet """
                for ii in range(3):
                    password = getpwd() #Frågar om lösenord och gör så att man inte ser vad som skrivs in i terminalen
                    pass_check = sheet_obj.cell(row = xx, column = 3)

                    if password == pass_check.value: # .value gör infomrationen till en str
                        print('You had the right password!')

                        """ Gör ett klass objekt av användaren """
                        role = sheet_obj.cell(row = xx, column = 5)
                        row = xx
                        
                        if role.value == 'User':
                            print('logged in as User')
                            balance = sheet_obj.cell(row = xx, column = 4)
                            logged_in_user = User(user_n, pass_check.value, 
                                            role.value, row, int(balance.value)) 
                            return logged_in_user
                        
                        elif role.value == 'Underage_user':
                            print('logged in as Underage_user')
                            balance = sheet_obj.cell(row = xx, column = 4)
                            logged_in_user = Underage_user(user_n, pass_check.value, 
                                            role.value, row, int(balance.value))
                            return logged_in_user

                        elif role.value == 'Admin':
                            print('logged in as Admin')
                            logged_in_user = Admin(user_n, pass_check.value, role.value, row)
                            return logged_in_user
                        break
                break
        
        if user_n.value == user_name and user_surn.value == user_surname:
            break # Stänger kvarvarande försök att skriva in rätt lösen och anvnam
        
        elif i<2:
            print('\nUser does not exist try again!\n')
            time.sleep(0.5)

        else:
            print('\nUser does not exist!\n')
            time.sleep(0.5)



""" Funktion som byter lösenord på användaren. Kan användas av admin och user """
def change_password(user):
    path = "./Employees.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    while True:
        new_password = str(input('New password: '))
        new_password_check = str(input('New password: '))

        if new_password == new_password_check:
            sheet_obj.cell(row=user.return_row(), column =3).value = f'{new_password}'
            wb_obj.save(r"./Employees.xlsx") 
            print('Your password is now changed!')
            time.sleep(1)
            break

        else:
            print('Password is incorrect try again!')



""" Admin funktioner och loopar """
""" Skapar ny användare """
def create_new_user():
    path = "./Employees.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_ro = sheet_obj.max_row

    while True:
        try:
            print() # För snygghetens skull i terminalen
            new_name = str(input('First name: '))
            new_surname = str(input('Surname: '))
            role = int(input('\nUser type:' +
                                '\nAdmin (1)' + 
                                '\nUser (2)'+ 
                                '\nUnderage_user (3)' +
                                '\nAnswer: '))
            new_user_password = str(input('Password: '))
            new_user_password_check = str(input('Password check: '))

            if new_user_password == new_user_password_check and role <=3 and role > 0:
                print('\nIf the following is true:' + 
                        '\n    First name: ', new_name +
                        '\n    Surname: ', new_surname +
                        '\n    Password: ', new_user_password +
                        '\n    User type: ', role)

                ans_2 = int(input('\nPress (1) if not press (2): '))

                if ans_2 == 1:
                    sheet_obj.cell(row=max_ro+1, column =1).value = f'{new_name}' # Skriver in värdet i excel
                    sheet_obj.cell(row=max_ro+1, column =2).value = f'{new_surname}'
                    sheet_obj.cell(row=max_ro+1, column =3).value = f'{new_user_password}'

                    while True:
                        try:
                            if role == 1:
                                sheet_obj.cell(row=max_ro+1, column =5).value = 'Admin'
                                break

                            elif role == 2:
                                sheet_obj.cell(row=max_ro+1, column =5).value = 'User'
                                money_ans = int(input('How much money does the preson have: '))
                                sheet_obj.cell(row=max_ro+1, column =4).value = f'{money_ans}'
                                break

                            elif role == 3:
                                sheet_obj.cell(row=max_ro+1, column =5).value = 'Underage_user'
                                money_ans = int(input('How much money does the preson have: '))
                                sheet_obj.cell(row=max_ro+1, column =4).value = f'{money_ans}'
                                break

                        except:
                            print('Answer is not valid try again!')

                    wb_obj.save(r"./Employees.xlsx") # Sparar informationen som skrivits in i excel filen
                    break
                
                elif ans_2 == 2:
                    print('Try again!')
            
            else:
                print('\nPassword is incorrect try again!')

        except:
            print('\nAnswer is not valid try again!')



""" Huvud funktion för admin """
def admin_loop(user):
    while True:
        try:
            move_ans = int(input('\n<------------What do you whant to do?' + 
                            '------------>' + 
                            '\nCreate new user (1)' + 
                            '\nChange password (2)' + 
                            '\nFind user information (3)' + 
                            '\nLoggout (4)' + 
                            '\nAnswer: '))

            if move_ans == 1:
                create_new_user()

            elif move_ans == 2:
                change_password(user)
            
            elif move_ans == 3: # Kollar efter en användare som man sedan kan få information om                
                path = "./Employees.xlsx"
                wb_obj = openpyxl.load_workbook(path)
                sheet_obj = wb_obj.active
                max_ro = sheet_obj.max_row

                users_name = str(input('\nUsers name: '))
                users_surname = str(input('Users surname: '))
                exist = False # Håller koll på om användaren finns eller inte

                for ro in range(1, max_ro+1):
                    user_na = sheet_obj.cell(row = ro, column = 1)
                    user_surna = sheet_obj.cell(row = ro, column = 2)

                    if user_na.value == users_name and user_surna.value == users_surname:
                        exist = True
                        ans_3 = int(input('\nWhat do you whant to know?' +
                                        '\nPassword (1)' + 
                                        '\nAccount balance (2)' +
                                        '\nUser type (3)' +
                                        '\nNothing (4)' +
                                        '\nAnswer: '))
                        if ans_3 == 1:
                            print('Password: ' ,sheet_obj.cell(row = ro, column = 3).value)
                            break
                        
                        elif ans_3 == 2:
                            print('Account balance: ' ,sheet_obj.cell(row = ro, column = 4).value)
                            break

                        elif ans_3 == 3:
                            print('User type: ' ,sheet_obj.cell(row = ro, column = 5).value)
                            break
                        
                        elif ans_3 == 4:
                            print('Ok then!')
                            break
                    
                if exist == False:
                    print('User does not exist!')

            elif move_ans == 4:
                print('Logging out!')
                time.sleep(1)
                break

        except:
            print('Answer is not valid try again!')



""" User loop """
def user_loop(user):
    while True:
        try:
            move_ans = int(input('\n<------------What do you whant to do?' +
                            '------------>' + 
                            '\nMoney money money money... MONEY!(In falsetto) (1)' + 
                            '\nChange password (2)' + 
                            '\nLoggout (3)' + 
                            '\nAnswer: '))

            if move_ans == 1:
                try:
                    print('\nThere are: ' + str(user.check_account_balance()) + # returnerar värdet som sattes in i kalssen
                                    ' Swedish kronor in the acccount.')
                    time.sleep(0.5)
                
                except:
                    print('Sum-ting-wong!')

            elif move_ans == 2:
                change_password(user)
                
            elif move_ans == 3:
                print('Logging out!')
                time.sleep(1)
                break

        except:
            print('Answer is not valid try again!')



""" Underage user loop. Har minst rättigheter. Kan bara kolla pengar """
def underage_user_loop(user):
    while True:
        try:
            move_ans = int(input('\n<------------What do you whant to do?' +
                            '------------>' + 
                            '\nMoney money money money... MONEY!(I falsett) (1)' + 
                            '\nLoggout (2)' + 
                            '\nAnswer: '))

            if move_ans == 1:
                print('\nThere are: ' + str(user.check_account_balance()) +
                                ' Swedish kronor in the acccount.')
                time.sleep(0.5)

            elif move_ans == 2:
                print('Logging out!')
                time.sleep(1)
                break

        except:
            print('Answer is not valid try again!')



""" API grej """
def get_chuck():
    chuck_req = requests.get('https://api.chucknorris.io/jokes/random') # Hämtar in information från API:n
    chuck = chuck_req.json()
    print(chuck['value'] + '\n') # Skriver ut information från API:n



""" Main Loop """
if __name__ == '__main__':
    while True:
        try:
            main_ans = int(input('\n<------------Main menu------------>'+
                            '\nLoggin (1)' +
                            '\nQuit program (2)' + 
                            '\nAnswer: '))

            if main_ans == 1:
                user = login()

                if user.return_role() == 'User':
                    user_loop(user)
                
                elif user.return_role() == 'Underage_user':
                    underage_user_loop(user)

                elif user.return_role() == 'Admin':
                    admin_loop(user)
            
            elif main_ans == 2:
                print('\nGoodbye!')
                print('Chucknorris joke:') 
                get_chuck()
                # Jag vet att den är in lallad men varför skulle jag 
                # behöva en api till det här programmet från första början
                break

        except:
            print('Answer is not valid try again!')
            time.sleep(1)




